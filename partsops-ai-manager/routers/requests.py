from __future__ import annotations

import asyncio
import json
import uuid
from datetime import datetime, timezone
from typing import Any, Optional, List, Dict
from fastapi import APIRouter, BackgroundTasks, Depends, Header, HTTPException, Body, File, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field as PydanticField
from sqlmodel import Session, select, desc

from database import get_session
from rbac import get_privileged_tenant, get_current_tenant, get_current_principal, CurrentPrincipal
from services.request_service import RequestService
from app.automation.storage import storage
from app.automation.rate_limiter import rate_limiter
from models import UploadArtifact, EventType
from event_store import emit_event
from fastapi import Request
from services.pipeline_runs import (
    TERMINAL_STATUSES,
    get_pipeline_run,
    list_run_events,
    run_once,
    serialize_run,
    start_pipeline_run,
)

# Import the new agent orchestrator
from app.agents import AgentOrchestrator, create_orchestrator

# Rate limit config (env-overridable)
import os as _os
_RATE_LIMIT = int(_os.getenv("PARTSOPS_INTAKE_RATE_LIMIT", "10"))
_RATE_WINDOW = int(_os.getenv("PARTSOPS_INTAKE_RATE_WINDOW", "60"))


def _rate_limit(request: Request, tenant_id: str):
    key = f"intake:{tenant_id}:{request.client.host if request.client else 'unknown'}"
    allowed, retry_after = rate_limiter.allow(key, _RATE_LIMIT, _RATE_WINDOW)
    if not allowed:
        raise HTTPException(
            status_code=429,
            detail=f"Rate limit exceeded. Try again in {retry_after} seconds.",
            headers={"Retry-After": str(retry_after)},
        )


router = APIRouter(prefix="/api", tags=["Requests & Attachments"])


class RawRequestPayload(BaseModel):
    source: str
    text: str
    customer_name: str = "Unknown"
    customer_phone: Optional[str] = None
    customer_email: Optional[str] = None
    vehicle_vin: Optional[str] = None
    priority: str = "normal"


class ManualCorrectionPayload(BaseModel):
    source_text: str
    corrected_parts_json: str
    correction_reason_tags: list[str] = []
    corrected_vehicle_json: Optional[str] = None


class MatchSelectionPayload(BaseModel):
    part_name: str = PydanticField(min_length=1)
    offer: dict[str, Any]


class WorkspaceActionPayload(BaseModel):
    target_state: Optional[str] = None
    reason: str = ""
    part_name: Optional[str] = None
    offer: Optional[dict[str, Any]] = None


class PipelineRunPayload(BaseModel):
    requested_lane: Optional[str] = PydanticField(default=None, max_length=64)


class ImportFromArtifactPayload(BaseModel):
    artifact_id: str
    source: str = "FILE_UPLOAD"
    customer_name: str = "File Upload Client"
    priority: str = "normal"


def _utcnow() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None)


def _require_action_permission(action: str, principal: CurrentPrincipal) -> None:
    permissions = RequestService._role_permissions(principal.role)
    required = {
        "select_offer": "can_manage_matching",
        "create_invoice": "can_create_invoice",
        "sync_erp": "can_sync_erp",
        "retry_pipeline": "can_retry_pipeline",
    }.get(action, "can_manage_matching")
    if not permissions.get(required, False):
        raise HTTPException(status_code=403, detail=f"Action '{action}' is not permitted for role {principal.role}")


def _require_transition_permission(target_state: str, principal: CurrentPrincipal) -> None:
    if target_state in RequestService.INTERNAL_ONLY_STATES:
        raise HTTPException(
            status_code=409,
            detail={"code": "INTERNAL_TRANSITION_ONLY", "target_state": target_state},
        )
    permissions = RequestService._role_permissions(principal.role)
    required = "can_manage_matching"
    if target_state == "APPROVED":
        required = "can_approve_pricing"
    elif target_state == "ERP_SYNCING":
        required = "can_sync_erp"
    if not permissions.get(required, False):
        raise HTTPException(status_code=403, detail=f"Transition to {target_state} is not permitted for role {principal.role}")


def _require_current_version(session: Session, request_id: str, tenant_id: str, version: Optional[str]) -> None:
    if not version:
        return
    request = RequestService.get_request(session, request_id, tenant_id)
    current_version = request.updated_at.isoformat() if request.updated_at else ""
    if version != current_version:
        raise HTTPException(status_code=409, detail={
            "code": "REQUEST_VERSION_CONFLICT",
            "current_version": current_version,
        })


@router.get("/session")
def get_session_identity(principal: CurrentPrincipal = Depends(get_current_principal)):
    """Expose the authenticated server principal; the UI must not infer a role."""
    return {
        "tenant_id": principal.tenant_id,
        "role": principal.role,
        "authenticated": principal.authenticated,
        "auth_mode": principal.auth_mode,
        "permissions": RequestService._role_permissions(principal.role),
    }


@router.post("/attachments/upload", status_code=201)
async def upload_attachment(
    file: UploadFile = File(...),
    tenant_id: str = Depends(get_current_tenant),
    principal: CurrentPrincipal = Depends(get_current_principal),
    request_id: Optional[str] = Header(None),
    session: Session = Depends(get_session),
):
    try:
        artifact_id = f"art_{uuid.uuid4().hex[:12]}"
        stored_path, safe_filename, size_bytes = storage.save_file(
            tenant_id=tenant_id,
            artifact_id=artifact_id,
            file_obj=file.file,
            original_filename=file.filename,
        )
        artifact = UploadArtifact(
            artifact_id=artifact_id,
            tenant_id=tenant_id,
            request_id=request_id,
            original_filename=file.filename,
            safe_filename=safe_filename,
            stored_path=stored_path,
            content_type=file.content_type,
            size_bytes=size_bytes,
            sha256=storage.calculate_sha256(stored_path),
            uploaded_by=principal.tenant_id,
            status="stored",
            created_at=_utcnow(),
        )
        session.add(artifact)
        session.commit()
        session.refresh(artifact)
        if request_id:
            emit_event(
                session=session,
                request_id=request_id,
                event_type=EventType.DOCUMENT_PARSED,
                actor_type="user",
                actor_id=principal.tenant_id,
                payload={"artifact_id": artifact_id, "filename": file.filename},
                tenant_id=tenant_id,
            )
        return {
            "status": "success",
            "artifact_id": artifact.artifact_id,
            "stored_path": artifact.stored_path,
            "sha256": artifact.sha256,
        }
    except Exception as exc:  # pragma: no cover - defensive
        if "stored_path" in locals():
            storage.delete_file(stored_path)
        raise HTTPException(status_code=500, detail=f"Upload failed: {exc}") from exc


@router.get("/requests")
def get_requests(
    session: Session = Depends(get_session),
    tenant_id: str = Depends(get_privileged_tenant),
    principal: CurrentPrincipal = Depends(get_current_principal),
):
    items = []
    for request in RequestService.get_requests(session, tenant_id):
        workspace = RequestService.get_workspace(session, request.request_id, tenant_id, principal.role)
        actions = workspace["allowed_actions"]
        targets = [action["target_state"] for action in actions if action.get("kind") == "transition" and action.get("target_state")]
        items.append({
            **request.model_dump(),
            "allowed_targets": targets,
            "allowed_actions": actions,
            "recommended_action": actions[0] if actions else None,
            "version": request.updated_at.isoformat() if request.updated_at else None,
            "is_blocked": request.status in {
                "FAILED", "ERP_SYNC_FAILED", "SUPPLIER_ISSUE", "NEEDS_CLARIFICATION",
                "NEEDS_MANUAL_PARSE", "CLIENT_REJECTED", "EXPIRED",
            },
        })
    return items


@router.post("/requests/import-from-artifact")
async def import_from_artifact(
    payload: ImportFromArtifactPayload,
    session: Session = Depends(get_session),
    tenant_id: str = Depends(get_privileged_tenant),
    request: Request = None,
):
    _rate_limit(request, tenant_id)
    """Import a request from a previously uploaded artifact (file)."""
    # Get the artifact
    artifact = session.exec(
        select(UploadArtifact).where(
            UploadArtifact.artifact_id == payload.artifact_id,
            UploadArtifact.tenant_id == tenant_id
        )
    ).first()
    
    if not artifact:
        raise HTTPException(status_code=404, detail="Artifact not found")
    
    if artifact.status != "stored":
        raise HTTPException(status_code=400, detail=f"Artifact not ready for import (status: {artifact.status})")
    
    file_type = artifact.content_type or ""
    filename = artifact.original_filename
    stored_path = artifact.stored_path
    
    def _do_import() -> dict:
        from agents import process_intake_request
        from services.supplier_service import _parse_supplier_table_file, _extract_supplier_table_rows
        from event_store import emit_event, EventType
        
        raw_rows, file_type = _parse_supplier_table_file(stored_path, filename, file_type)
        normalized_rows, mapped_columns, validation_summary = _extract_supplier_table_rows(raw_rows)
        
        text_parts = []
        for row in normalized_rows:
            part_name = row.get("part_name", "")
            if part_name:
                oem = row.get("oem_number", "")
                brand = row.get("brand", "")
                qty = row.get("stock_qty", 1)
                text_parts.append(f"{part_name} {oem} {brand} x{qty}")
        
        text_content = "\n".join(text_parts) if text_parts else f"File upload: {filename}"
        
        intake_result = process_intake_request(
            text=text_content,
            priority=payload.priority,
        )
        
        request_payload = {
            "source": payload.source,
            "text": text_content,
            "customer_name": payload.customer_name,
            "priority": payload.priority,
        }
        
        if intake_result.get("vehicle_make"):
            request_payload["vehicle_vin"] = intake_result.get("vehicle_vin") or ""
        
        new_request = RequestService.create_request(tenant_id, request_payload, None)
        return {
            "text_content": text_content,
            "intake_result": intake_result,
            "new_request": new_request,
        }
    
    try:
        result = await asyncio.to_thread(_do_import)
        new_request = result["new_request"]
        
        emit_event(
            session=session,
            request_id=new_request["request_id"],
            event_type=EventType.DOCUMENT_PARSED,
            actor_type="user",
            actor_id="file_import",
            payload={"artifact_id": artifact.artifact_id, "filename": filename},
            tenant_id=tenant_id,
        )
        
        artifact.status = "attached"
        artifact.request_id = new_request["request_id"]
        session.add(artifact)
        session.commit()
        
        return new_request
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Import failed: {e}")


@router.post("/requests")
async def create_request(
    payload: RawRequestPayload,
    session: Session = Depends(get_session),
    x_idempotency_key: Optional[str] = Header(default=None),
    tenant_id: str = Depends(get_privileged_tenant),
    request: Request = None,
):
    _rate_limit(request, tenant_id)
    return await asyncio.to_thread(
        RequestService.create_request,
        tenant_id,
        payload.model_dump(),
        x_idempotency_key,
    )


@router.get("/requests/{request_id}")
def get_request(
    request_id: str,
    session: Session = Depends(get_session),
    tenant_id: str = Depends(get_privileged_tenant),
):
    return RequestService.get_request(session, request_id, tenant_id)


@router.get("/requests/{request_id}/workspace")
def get_request_workspace(
    request_id: str,
    session: Session = Depends(get_session),
    tenant_id: str = Depends(get_privileged_tenant),
    principal: CurrentPrincipal = Depends(get_current_principal),
):
    return RequestService.get_workspace(session, request_id, tenant_id, principal.role)


@router.post("/requests/{request_id}/transition")
def transition_state(
    request_id: str,
    body: dict[str, Any] = Body(...),
    session: Session = Depends(get_session),
    tenant_id: str = Depends(get_privileged_tenant),
    principal: CurrentPrincipal = Depends(get_current_principal),
    x_request_version: Optional[str] = Header(default=None),
):
    target_state = body.get("target_state")
    reason = body.get("reason", "")
    if not target_state:
        raise HTTPException(status_code=400, detail="target_state is required")
    _require_transition_permission(target_state, principal)
    _require_current_version(session, request_id, tenant_id, x_request_version)
    transition = RequestService.transition_state(
        session, request_id, tenant_id, target_state, reason, actor_id=f"operator:{principal.role}"
    )
    return {"transition": transition, "workspace": RequestService.get_workspace(session, request_id, tenant_id, principal.role)}


@router.post("/requests/{request_id}/actions/{action}")
def execute_workspace_action(
    request_id: str,
    action: str,
    payload: WorkspaceActionPayload,
    session: Session = Depends(get_session),
    tenant_id: str = Depends(get_privileged_tenant),
    principal: CurrentPrincipal = Depends(get_current_principal),
    x_request_version: Optional[str] = Header(default=None),
):
    """Canonical mutation surface for the operator workspace."""
    _require_current_version(session, request_id, tenant_id, x_request_version)
    if action == "transition":
        if not payload.target_state:
            raise HTTPException(status_code=400, detail="target_state is required")
        _require_transition_permission(payload.target_state, principal)
        RequestService.transition_state(
            session, request_id, tenant_id, payload.target_state, payload.reason,
            actor_id=f"operator:{principal.role}",
        )
    elif action == "select_offer":
        _require_action_permission(action, principal)
        if not payload.part_name or not payload.offer:
            raise HTTPException(status_code=400, detail="part_name and offer are required")
        RequestService.select_match(
            session, request_id, tenant_id, payload.part_name, payload.offer, actor_id=f"operator:{principal.role}"
        )
    elif action == "create_invoice":
        _require_action_permission(action, principal)
        RequestService.generate_invoice(session, request_id, tenant_id)
    else:
        raise HTTPException(status_code=404, detail=f"Unsupported workspace action: {action}")
    return RequestService.get_workspace(session, request_id, tenant_id, principal.role)


@router.post("/requests/{request_id}/correction")
def create_manual_correction(
    request_id: str,
    payload: ManualCorrectionPayload,
    session: Session = Depends(get_session),
    tenant_id: str = Depends(get_privileged_tenant),
):
    return RequestService.create_manual_correction(session, request_id, tenant_id, payload.model_dump())


@router.post("/requests/{request_id}/pipeline-runs", status_code=202)
def start_request_pipeline_run(
    request_id: str,
    payload: PipelineRunPayload,
    session: Session = Depends(get_session),
    tenant_id: str = Depends(get_privileged_tenant),
    principal: CurrentPrincipal = Depends(get_current_principal),
):
    run, idempotent = start_pipeline_run(
        session,
        request_id=request_id,
        tenant_id=tenant_id,
        requested_by=f"operator:{principal.role}",
        requested_lane=payload.requested_lane,
    )
    response = serialize_run(run, idempotent=idempotent)
    if idempotent:
        from fastapi.responses import JSONResponse
        return JSONResponse(status_code=200, content=response)
    return response


@router.get("/requests/{request_id}/pipeline-runs/{run_id}")
def get_request_pipeline_run(
    request_id: str,
    run_id: str,
    session: Session = Depends(get_session),
    tenant_id: str = Depends(get_privileged_tenant),
):
    return serialize_run(get_pipeline_run(session, request_id=request_id, run_id=run_id, tenant_id=tenant_id))


@router.get("/requests/{request_id}/pipeline-runs/{run_id}/events")
async def stream_request_pipeline_run_events(
    request_id: str,
    run_id: str,
    after: int = 0,
    session: Session = Depends(get_session),
    tenant_id: str = Depends(get_privileged_tenant),
):
    # Resolve once before opening the stream so access failures are ordinary HTTP errors.
    get_pipeline_run(session, request_id=request_id, run_id=run_id, tenant_id=tenant_id)

    async def event_stream():
        last_sequence = after
        while True:
            events = list_run_events(session, run_id=run_id, tenant_id=tenant_id, after=last_sequence)
            for event in events:
                last_sequence = event.sequence
                try:
                    payload = json.loads(event.payload_json or "{}")
                except json.JSONDecodeError:
                    payload = {}
                data = json.dumps(
                    {
                        "sequence": event.sequence,
                        "type": event.event_type,
                        "phase": event.phase,
                        "message": event.message,
                        "payload": payload,
                    },
                    ensure_ascii=False,
                )
                yield f"id: {event.sequence}\nevent: {event.event_type}\ndata: {data}\n\n"
            run = get_pipeline_run(session, request_id=request_id, run_id=run_id, tenant_id=tenant_id)
            if run.status in TERMINAL_STATUSES:
                return
            yield ": keepalive\n\n"
            await asyncio.sleep(0.5)

    return StreamingResponse(event_stream(), media_type="text/event-stream", headers={"Cache-Control": "no-cache"})


@router.get("/requests/{request_id}/matches")
def get_selected_matches(
    request_id: str,
    session: Session = Depends(get_session),
    tenant_id: str = Depends(get_privileged_tenant),
):
    return RequestService.get_selected_matches(session, request_id, tenant_id)


@router.post("/requests/{request_id}/matches")
def select_match(
    request_id: str,
    payload: MatchSelectionPayload,
    session: Session = Depends(get_session),
    tenant_id: str = Depends(get_privileged_tenant),
    principal: CurrentPrincipal = Depends(get_current_principal),
):
    _require_action_permission("select_offer", principal)
    return RequestService.select_match(
        session,
        request_id,
        tenant_id,
        payload.part_name,
        payload.offer,
        f"operator:{principal.role}",
    )


@router.get("/requests/{request_id}/gates")
def get_request_gates(
    request_id: str,
    session: Session = Depends(get_session),
    tenant_id: str = Depends(get_privileged_tenant),
):
    return RequestService.get_request_gates(session, request_id, tenant_id)


@router.get("/requests/{request_id}/events")
def get_request_events(
    request_id: str,
    session: Session = Depends(get_session),
    tenant_id: str = Depends(get_privileged_tenant),
):
    return RequestService.get_request_events(session, request_id, tenant_id)


@router.get("/requests/{request_id}/audit")
def audit_event_chain(
    request_id: str,
    session: Session = Depends(get_session),
    tenant_id: str = Depends(get_privileged_tenant),
):
    return RequestService.audit_event_chain(session, request_id, tenant_id)


# ──────────────────────────────────────────────
# MULTI-AGENT PIPELINE ENDPOINTS
# ──────────────────────────────────────────────

class PipelineRequestPayload(BaseModel):
    """Payload for running the full multi-agent pipeline"""
    source: str  # telegram, email, crm, web, manual, api
    text: str
    customer_name: Optional[str] = None
    customer_phone: Optional[str] = None
    customer_email: Optional[str] = None
    customer_erp_id: Optional[str] = None
    vehicle_vin: Optional[str] = None
    vehicle_make: Optional[str] = None
    vehicle_model: Optional[str] = None
    vehicle_year: Optional[int] = None
    vehicle_generation: Optional[str] = None
    vehicle_engine: Optional[str] = None
    parts_data: Optional[List[Dict[str, Any]]] = None
    metadata: Optional[Dict[str, Any]] = None
    priority: str = "normal"


class PipelineContinuePayload(BaseModel):
    """Payload for continuing a pipeline from a specific stage"""
    start_from: str = "processing"  # intake, processing, delivery, reporting


@router.post("/pipeline/run")
def run_full_pipeline(
    payload: PipelineRequestPayload,
    session: Session = Depends(get_session),
    tenant_id: str = Depends(get_privileged_tenant),
):
    """
    Run the complete multi-agent order processing pipeline.
    
    Flow: Intake → Processing → Delivery → Reporting
    
    This endpoint:
    1. Creates/validates the order (Intake Agent)
    2. Matches parts and calculates pricing (Processing Agent)
    3. Generates approval document
    4. Sends to client via appropriate channel (Delivery Agent)
    5. Reports results to operators (Reporting Agent)
    """
    orchestrator = create_orchestrator(tenant_id=tenant_id)
    
    
    try:
        result = orchestrator.run_pipeline(
            source=payload.source,
            raw_input=payload.text,
            customer_data={
                "name": payload.customer_name,
                "phone": payload.customer_phone,
                "email": payload.customer_email,
                "erp_id": payload.customer_erp_id,
            },
            vehicle_data={
                "vin": payload.vehicle_vin,
                "make": payload.vehicle_make,
                "model": payload.vehicle_model,
                "year": payload.vehicle_year,
                "generation": payload.vehicle_generation,
                "engine": payload.vehicle_engine,
            },
            parts_data=payload.parts_data,
            metadata=payload.metadata or {},
            priority=payload.priority,
        )
        
        return {
            "success": result.success,
            "request_id": result.request_id,
            "order_id": result.order_id,
            "phases": {k: v.to_dict() for k, v in result.phases.items()},
            "errors": result.errors,
            "warnings": result.warnings,
            "correlation_id": result.correlation_id,
            "total_time_ms": result.total_time_ms,
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Pipeline execution failed: {e}")


@router.post("/pipeline/continue/{request_id}")
def continue_pipeline(
    request_id: str,
    payload: PipelineContinuePayload,
    session: Session = Depends(get_session),
    tenant_id: str = Depends(get_privileged_tenant),
):
    """
    Continue a pipeline from a specific stage (for retries).
    
    Args:
        request_id: The request to continue
        start_from: Which agent to start from (intake, processing, delivery, reporting)
    """
    from app.agents import AgentType
    
    try:
        start_agent = AgentType(payload.start_from)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid start_from: {payload.start_from}")
    
    orchestrator = create_orchestrator(tenant_id=tenant_id)
    
    try:
        result = orchestrator.continue_pipeline(
            request_id=request_id,
            start_from=start_agent,
        )
        
        return {
            "success": result.success,
            "request_id": result.request_id,
            "order_id": result.order_id,
            "phases": {k: v.to_dict() for k, v in result.phases.items()},
            "errors": result.errors,
            "warnings": result.warnings,
            "correlation_id": result.correlation_id,
            "total_time_ms": result.total_time_ms,
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Pipeline continuation failed: {e}")


@router.get("/pipeline/status/{request_id}")
def get_pipeline_status(
    request_id: str,
    session: Session = Depends(get_session),
    tenant_id: str = Depends(get_privileged_tenant),
):
    """Get the current status of a request in the pipeline"""
    from models import PartRequest
    
    request = session.exec(
        select(PartRequest).where(
            PartRequest.request_id == request_id,
            PartRequest.tenant_id == tenant_id
        )
    ).first()
    
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    
    return {
        "request_id": request.request_id,
        "status": request.status,
        "source": request.source,
        "priority": request.priority,
        "customer_name": request.customer_name,
        "vehicle": {
            "make": request.vehicle_make,
            "model": request.vehicle_model,
            "year": request.vehicle_year,
            "vin": request.vehicle_vin_masked,
        },
        "parts_count": len(request.parts_json) if request.parts_json else 0,
        "pricing_total": None,  # Would need to parse pricing_evidence_json
        "original_ref": request.raw_input_ref,
        "created_at": request.created_at.isoformat() if request.created_at else None,
        "updated_at": request.updated_at.isoformat() if request.updated_at else None,
    }


# ──────────────────────────────────────────────
# DELIVERY ENDPOINTS (for InvoicePreview component)
# ──────────────────────────────────────────────

@router.get("/delivery/status/{request_id}")
def get_delivery_status(
    request_id: str,
    session: Session = Depends(get_session),
    tenant_id: str = Depends(get_privileged_tenant),
):
    """Get delivery status/logs for a request"""
    from models import OutboundMessage
    from suppliers import Invoice
    from sqlmodel import select
    
    # Get outbound messages for this request
    messages = session.exec(
        select(OutboundMessage).where(
            OutboundMessage.request_id == request_id,
            OutboundMessage.tenant_id == tenant_id,
        ).order_by(OutboundMessage.created_at.desc())
    ).all()
    
    return [
        {
            "message_id": m.id,
            "channel": m.channel,
            "recipient": m.recipient,
            "status": m.status,
            "attempts": m.attempts,
            "last_error": m.last_error,
            "sent_at": m.sent_at.isoformat() if m.sent_at else None,
            "created_at": m.created_at.isoformat(),
        }
        for m in messages
    ]


@router.get("/delivery/invoice/{request_id}/pdf")
def get_invoice_pdf(
    request_id: str,
    session: Session = Depends(get_session),
    tenant_id: str = Depends(get_privileged_tenant),
):
    """Get PDF for invoice linked to request"""
    from suppliers import Invoice
    from delivery import InvoicePDFGenerator
    from fastapi.responses import Response
    from sqlmodel import select
    
    # Find invoice for this request
    invoice = session.exec(
        select(Invoice).where(
            Invoice.request_id == request_id,
            Invoice.tenant_id == tenant_id,
        )
    ).first()
    
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found for this request")
    
    # Generate PDF
    pdf_bytes = InvoicePDFGenerator.generate(invoice)
    
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'inline; filename="Invoice-{invoice.invoice_number}.pdf"'
        }
    )


@router.post("/delivery/send/{request_id}")
def send_invoice(
    request_id: str,
    body: dict,
    session: Session = Depends(get_session),
    tenant_id: str = Depends(get_privileged_tenant),
):
    """Send invoice via specified channel"""
    from suppliers import Invoice
    from delivery import EmailAdapter, TelegramAdapter
    from sqlmodel import select
    
    channel = body.get("channel", "email")
    recipient = body.get("recipient", "")
    dry_run = body.get("dry_run", False)
    
    if not recipient:
        raise HTTPException(status_code=400, detail="Recipient is required")
    
    # Find invoice for this request
    invoice = session.exec(
        select(Invoice).where(
            Invoice.request_id == request_id,
            Invoice.tenant_id == tenant_id,
        )
    ).first()
    
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found for this request")
    
    try:
        if channel == "email":
            result = EmailAdapter.send_invoice(
                invoice=invoice,
                recipient_email=recipient,
                session=session,
                tenant_id=tenant_id,
                dry_run=dry_run,
            )
        elif channel == "telegram":
            result = TelegramAdapter.send_invoice_preview(
                invoice=invoice,
                chat_id=recipient,
                session=session,
                tenant_id=tenant_id,
                dry_run=dry_run,
            )
        else:
            raise HTTPException(status_code=400, detail=f"Unsupported channel: {channel}")
        
        return {
            "success": result.status == "sent",
            "message_id": result.id,
            "status": result.status,
            "channel": result.channel,
            "recipient": result.recipient,
            "error": result.last_error,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to send: {e}")


# ──────────────────────────────────────────────
# APPROVAL WORKFLOW ENDPOINTS
# ──────────────────────────────────────────────

class ApprovalActionPayload(BaseModel):
    """Payload for approve/reject actions"""
    action: str  # "approve" or "reject"
    comment: Optional[str] = None
    actor_id: str = "admin"


@router.post("/requests/{request_id}/approve")
def approve_request(
    request_id: str,
    payload: ApprovalActionPayload,
    session: Session = Depends(get_session),
    tenant_id: str = Depends(get_privileged_tenant),
):
    """Approve or reject a request awaiting approval"""
    from models import PartRequest, ApprovalTicket, RequestState, EventType
    from sqlmodel import select
    
    if payload.action not in ("approve", "reject"):
        raise HTTPException(status_code=400, detail="Action must be 'approve' or 'reject'")
    
    request = session.exec(
        select(PartRequest).where(
            PartRequest.request_id == request_id,
            PartRequest.tenant_id == tenant_id
        )
    ).first()
    
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    
    if request.status != RequestState.READY_FOR_APPROVAL:
        raise HTTPException(status_code=400, detail=f"Request not in READY_FOR_APPROVAL state (current: {request.status})")
    
    # Find pending approval ticket
    ticket = session.exec(
        select(ApprovalTicket).where(
            ApprovalTicket.request_id == request_id,
            ApprovalTicket.tenant_id == tenant_id,
            ApprovalTicket.status == "pending"
        )
    ).first()
    
    if payload.action == "approve":
        # Approve the request
        request.status = RequestState.APPROVED
        request.updated_at = datetime.utcnow()
        session.add(request)
        
        # Update ticket
        if ticket:
            ticket.status = "approved"
            ticket.decided_by = payload.actor_id
            ticket.decided_at = datetime.utcnow()
            ticket.decision_note = payload.comment
            session.add(ticket)
        
        # Emit event
        emit_event(
            session=session,
            request_id=request_id,
            event_type=EventType.MANAGER_APPROVED,
            actor_type="user",
            actor_id=payload.actor_id,
            payload={"comment": payload.comment},
            tenant_id=tenant_id,
        )
        
        message = "Request approved successfully"
        
        # Continue pipeline after approval (delivery + reporting)
        try:
            from app.agents import create_orchestrator, AgentType
            orchestrator = create_orchestrator(tenant_id=tenant_id)
            continue_result = orchestrator.continue_pipeline(
                request_id=request_id,
                start_from=AgentType.DELIVERY,
            )
            if continue_result.success:
                delivery_success = continue_result.phases.get('delivery')
                delivery_success = delivery_success.success if delivery_success else False
                reporting_success = continue_result.phases.get('reporting')
                reporting_success = reporting_success.success if reporting_success else False
                message += f" | Pipeline continued: delivery={delivery_success}, reporting={reporting_success}"
            else:
                message += f" | Pipeline continuation failed: {continue_result.errors}"
        except Exception as e:
            message += f" | Pipeline continuation error: {str(e)}"
        
    else:
        # Reject the request
        request.status = RequestState.CLIENT_REJECTED
        request.updated_at = datetime.utcnow()
        session.add(request)
        
        # Update ticket
        if ticket:
            ticket.status = "rejected"
            ticket.decided_by = payload.actor_id
            ticket.decided_at = datetime.utcnow()
            ticket.decision_note = payload.comment
            session.add(ticket)
        
        # Emit event
        emit_event(
            session=session,
            request_id=request_id,
            event_type=EventType.MANAGER_REJECTED,
            actor_type="user",
            actor_id=payload.actor_id,
            payload={"comment": payload.comment, "reason": "rejected_by_manager"},
            tenant_id=tenant_id,
        )
        
        message = "Request rejected"
    
    session.commit()
    
    return {
        "success": True,
        "request_id": request_id,
        "new_status": request.status,
        "message": message,
    }


@router.get("/requests/{request_id}/approval-tickets")
def get_approval_tickets(
    request_id: str,
    session: Session = Depends(get_session),
    tenant_id: str = Depends(get_privileged_tenant),
):
    """Get approval tickets for a request"""
    from models import ApprovalTicket
    from sqlmodel import select
    
    tickets = session.exec(
        select(ApprovalTicket).where(
            ApprovalTicket.request_id == request_id,
            ApprovalTicket.tenant_id == tenant_id,
        ).order_by(ApprovalTicket.created_at.desc())
    ).all()
    
    return [
        {
            "ticket_id": t.ticket_id,
            "tool_name": t.tool_name,
            "reason": t.reason,
            "role_required": t.role_required,
            "status": t.status,
            "requested_by": t.requested_by,
            "decided_by": t.decided_by,
            "decided_at": t.decided_at.isoformat() if t.decided_at else None,
            "decision_note": t.decision_note,
            "created_at": t.created_at.isoformat(),
        }
        for t in tickets
    ]


# ──────────────────────────────────────────────
# CLIENT PORTAL ENDPOINTS (Public tracking)
# ──────────────────────────────────────────────

from pydantic import BaseModel

class ClientPortalViewPayload(BaseModel):
    """Public view of request for client portal"""
    tracking_token: str

class ClientPortalActionPayload(BaseModel):
    """Action from client (accept/reject)"""
    tracking_token: str
    action: str  # "accept" or "reject"
    reason: Optional[str] = None


@router.get("/client/track/{token}")
def client_track_request(
    token: str,
    session: Session = Depends(get_session),
    tenant_id: str = Depends(get_current_tenant),
):
    """Get public view of request via tracking token"""
    from client_portal import get_public_request_view
    
    view = get_public_request_view(token, session, tenant_id)
    if not view:
        raise HTTPException(status_code=404, detail="Request not found or token invalid")
    
    return view


@router.post("/client/track/{token}/accept")
def client_accept_offer(
    token: str,
    session: Session = Depends(get_session),
    tenant_id: str = Depends(get_current_tenant),
):
    """Client accepts the offer (SENT_TO_CLIENT -> PAID)"""
    from client_portal import accept_offer
    
    result = accept_offer(token, session, tenant_id)
    if not result["ok"]:
        raise HTTPException(status_code=400, detail=result["error"])
    
    return result


@router.post("/client/track/{token}/reject")
def client_reject_offer(
    token: str,
    body: dict,
    session: Session = Depends(get_session),
    tenant_id: str = Depends(get_current_tenant),
):
    """Client rejects the offer (SENT_TO_CLIENT -> CLIENT_REJECTED)"""
    from client_portal import reject_offer
    
    reason = body.get("reason", "No reason provided")
    result = reject_offer(token, reason, session, tenant_id)
    if not result["ok"]:
        raise HTTPException(status_code=400, detail=result["error"])
    
    return result


@router.post("/requests/{request_id}/generate-tracking-token")
def generate_tracking_token_endpoint(
    request_id: str,
    session: Session = Depends(get_session),
    tenant_id: str = Depends(get_privileged_tenant),
):
    """Generate tracking token for client portal access"""
    from client_portal import create_tracking_token
    from models import PartRequest
    from sqlmodel import select
    from datetime import datetime, timedelta
    
    request = session.exec(
        select(PartRequest).where(
            PartRequest.request_id == request_id,
            PartRequest.tenant_id == tenant_id
        )
    ).first()
    
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    
    token = create_tracking_token(
        request_id=request_id,
        tenant_id=tenant_id,
        expires_hours=72,
    )
    
    # Store token in request
    request.tracking_token = token
    request.tracking_token_expires_at = datetime.utcnow() + timedelta(hours=72)
    session.add(request)
    session.commit()
    
    return {
        "success": True,
        "request_id": request_id,
        "tracking_token": token,
        "tracking_url": f"/client/track/{token}",
        "expires_at": request.tracking_token_expires_at.isoformat(),
    }


@router.get("/requests/{request_id}/export-excel")
def export_request_excel(
    request_id: str,
    session: Session = Depends(get_session),
    tenant_id: str = Depends(get_privileged_tenant),
):
    """Export request specifications and quotation details as a styled XLSX file"""
    import io
    import json
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from models import PartRequest

    req = session.exec(
        select(PartRequest).where(
            PartRequest.request_id == request_id,
            PartRequest.tenant_id == tenant_id
        )
    ).first()

    if not req:
        raise HTTPException(status_code=404, detail="Request not found")

    wb = Workbook()
    ws = wb.active
    ws.title = "Спецификация Заказа"
    ws.views.sheetView[0].showGridLines = True

    # Styles
    font_title = Font(name="Arial", size=14, bold=True, color="0F172A")
    font_subtitle = Font(name="Arial", size=10, italic=True, color="64748B")
    font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    font_data = Font(name="Arial", size=10, color="0F172A")
    font_total = Font(name="Arial", size=11, bold=True, color="0F172A")

    fill_header = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    fill_total = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    # Title Block
    ws.append(["PARTSOPS AI MANAGER — ИТОГОВЫЙ ОТЧЕТ ПО ЗАДАНИЮ И СМЕТА"])
    ws.cell(row=1, column=1).font = font_title
    ws.append([f"Сгенерировано: {datetime.utcnow().strftime('%d.%m.%Y %H:%M UTC')} | Система PartsOps AI v6.0"])
    ws.cell(row=2, column=1).font = font_subtitle
    ws.append([])

    # Metadata Block
    ws.append(["ID Запроса:", req.request_id, "", "Клиент:", req.customer_name or "Не указан"])
    ws.append(["Статус:", req.status, "", "Автомобиль:", f"{req.vehicle_make or ''} {req.vehicle_model or ''}".strip() or "Уточняется"])
    ws.append(["Приоритет:", req.priority.upper(), "", "VIN:", req.vehicle_vin_masked or "—"])
    ws.append([])

    # Table Header
    headers = [
        "№", "Артикул OEM", "Наименование детали", "Поставщик",
        "Наличие (шт)", "Срок (дн)", "Закупка (руб)", "Наценка (%)",
        "Цена клиенту (руб)", "Итого к оплате (руб)", "Match Score"
    ]
    ws.append(headers)
    header_row = ws.max_row

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Parse parts json
    parts_list = []
    try:
        if req.parts_json:
            parts_list = json.loads(req.parts_json)
    except Exception:
        parts_list = []
    if not isinstance(parts_list, list) or not parts_list:
        raise HTTPException(status_code=422, detail="Экспорт заблокирован: в заявке нет подтверждённых позиций")

    start_data_row = header_row + 1

    for idx, item in enumerate(parts_list, start=1):
        name = item.get("name") or item.get("part_name") or ""
        oem = item.get("oem") or item.get("oem_number") or item.get("article") or ""
        qty = int(item.get("quantity") or 0)
        supplier = item.get("supplier_name") or ""
        stock_qty = item.get("stock_qty")
        delivery_days = item.get("delivery_days")
        buy_price = item.get("price")
        margin_pct = item.get("margin_pct") or item.get("margin")
        client_price = item.get("client_price")
        row_total = item.get("line_total")
        match_score = item.get("match_score") or item.get("score") or ""

        row_data = [idx, oem, name, supplier, stock_qty, delivery_days, buy_price, margin_pct, client_price, row_total, match_score]
        ws.append(row_data)
        curr_row = ws.max_row

        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.font = font_data
            cell.border = thin_border
            if col_idx in (1, 5, 6, 8, 11):
                cell.alignment = Alignment(horizontal="center")
            elif col_idx in (7, 9, 10):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0.00 "₽"'

    end_data_row = ws.max_row
    # Summary Row
    ws.append(["Итого по спецификации", "", "", "", "", "", f"=SUM(G{start_data_row}:G{end_data_row})", "", f"=SUM(I{start_data_row}:I{end_data_row})", f"=SUM(J{start_data_row}:J{end_data_row})", ""])
    total_row = ws.max_row

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.font = font_total
        cell.fill = fill_total
        cell.border = thin_border
        if col_idx in (7, 9, 10):
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = '#,##0.00 "₽"'

    # Auto-adjust column widths for Sheet 1
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # ----------------------------------------------------
    # Sheet 2: Аналоги и Замены (Smart Analog Report)
    # ----------------------------------------------------
    from models import AnalogCandidate, ContractPosition
    ws2 = wb.create_sheet(title="Аналоги и Замены")
    ws2.views.sheetView[0].showGridLines = True

    ws2.append(["АНАЛИЗ АЛЬТЕРНАТИВ И ПОДБОР АНАЛОГОВ (SMART FALLBACK ENGINE)"])
    ws2.cell(row=1, column=1).font = font_title
    ws2.append([f"Сформировано по заявке: {req.request_id} | Проверка совместимости и рисков"])
    ws2.cell(row=2, column=1).font = font_subtitle
    ws2.append([])

    headers_ws2 = [
        "№", "Оригинальный OEM", "Статус оригинала", "Артикул аналога",
        "Бренд аналога", "Категория (Tier)", "Уровень риска", "Обоснование / Факторы риска", "Статус выборки"
    ]
    ws2.append(headers_ws2)
    header_row2 = ws2.max_row

    for col_idx in range(1, len(headers_ws2) + 1):
        cell = ws2.cell(row=header_row2, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Fills for Tiers
    fill_oes = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    font_oes = Font(name="Arial", size=10, bold=True, color="065F46")
    
    fill_prem = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    font_prem = Font(name="Arial", size=10, bold=True, color="1E40AF")

    fill_budg = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    font_budg = Font(name="Arial", size=10, bold=True, color="92400E")

    fill_spec = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    font_spec = Font(name="Arial", size=10, bold=True, color="991B1B")

    # Query analogs for this request
    contract_positions = session.exec(
        select(ContractPosition).where(
            ContractPosition.request_id == request_id,
            ContractPosition.tenant_id == tenant_id
        )
    ).all()

    analogs_list = []
    if contract_positions:
        pos_ids = [p.position_id for p in contract_positions]
        analogs_list = session.exec(
            select(AnalogCandidate).where(
                AnalogCandidate.position_id.in_(pos_ids),
                AnalogCandidate.tenant_id == tenant_id
            )
        ).all()

    if not analogs_list:
        # Honesty: no demo/sample rows when DB has no analogs
        ws2.append([
            "—",
            "—",
            "Нет аналогов в БД",
            "—",
            "—",
            "—",
            "—",
            "Запустите resolve-analogs / crawler; demo-строки отключены",
            "пусто",
        ])
        for c in range(1, 10):
            ws2.cell(row=ws2.max_row, column=c).border = thin_border
    else:
        for idx, analog in enumerate(analogs_list, start=1):
            tier_label = f"{analog.quality_tier} (Tier)"
            risk_label = f"{analog.risk_score}%"
            factors = analog.risk_factors_json or "Верифицированный кросс-номер"
            row2_data = [
                idx,
                analog.previous_article or "OEM Ref",
                "OEM Недоступен",
                analog.article,
                analog.brand,
                tier_label,
                risk_label,
                factors,
                analog.manual_review_status.upper()
            ]
            ws2.append(row2_data)
            row_idx = ws2.max_row
            tier_cell = ws2.cell(row=row_idx, column=6)
            
            if analog.quality_tier == "OES":
                tier_cell.fill = fill_oes
                tier_cell.font = font_oes
            elif analog.quality_tier == "PREMIUM_AFTERMARKET":
                tier_cell.fill = fill_prem
                tier_cell.font = font_prem
            elif analog.quality_tier == "BUDGET":
                tier_cell.fill = fill_budg
                tier_cell.font = font_budg
            else:
                tier_cell.fill = fill_spec
                tier_cell.font = font_spec

            for c in range(1, len(row2_data) + 1):
                ws2.cell(row=row_idx, column=c).border = thin_border

    # Auto-adjust column widths for Sheet 2
    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 14)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"partsops_report_{request_id}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
