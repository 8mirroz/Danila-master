from __future__ import annotations

import io
from typing import Literal

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlmodel import Session

from database import get_session
from rbac import CurrentPrincipal, get_current_principal, get_privileged_tenant
from services.quotes import get_quote, issue_quote, list_quotes

router = APIRouter(prefix="/api/quotes", tags=["Quotes"])


class IssueQuotePayload(BaseModel):
    request_id: str = Field(min_length=1)
    valid_for_days: int = Field(default=14, ge=1, le=90)


@router.get("")
def get_quotes(
    session: Session = Depends(get_session),
    organization_id: str = Depends(get_privileged_tenant),
):
    return list_quotes(session, organization_id=organization_id)


@router.post("", status_code=201)
def create_quote(
    payload: IssueQuotePayload,
    session: Session = Depends(get_session),
    organization_id: str = Depends(get_privileged_tenant),
    principal: CurrentPrincipal = Depends(get_current_principal),
):
    if principal.role not in {"admin", "finance"}:
        raise HTTPException(403, "Quote issuing requires finance or admin role")
    return issue_quote(
        session,
        organization_id=organization_id,
        request_id=payload.request_id,
        valid_for_days=payload.valid_for_days,
        created_by=principal.subject or f"operator:{principal.role}",
    )


@router.get("/{quote_id}")
def read_quote(
    quote_id: str,
    version: int | None = None,
    session: Session = Depends(get_session),
    organization_id: str = Depends(get_privileged_tenant),
):
    return get_quote(
        session, organization_id=organization_id, quote_id=quote_id, version=version
    )


@router.get("/{quote_id}/export/{format}")
def export_quote(
    quote_id: str,
    format: Literal["xlsx", "pdf"],
    version: int | None = None,
    session: Session = Depends(get_session),
    organization_id: str = Depends(get_privileged_tenant),
):
    quote = get_quote(
        session, organization_id=organization_id, quote_id=quote_id, version=version
    )
    if format == "xlsx":
        return _xlsx(quote)
    return _pdf(quote)


def _xlsx(quote: dict):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Quote"
    sheet.append(["PartsOps Quote", quote["quote_id"], f"Version {quote['version']}"])
    sheet.append(["Valid until", quote["valid_until"]])
    sheet.append([])
    sheet.append(["Part", "Supplier", "Unit price", "Quantity", "Line total"])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="123A78")
    for cell in sheet[4]:
        cell.font = Font(bold=True)
    for line in quote["pricing"].get("line_items", []):
        sheet.append(
            [
                line.get("part_name"),
                line.get("supplier_name") or line.get("supplier_id"),
                line.get("sale_price"),
                line.get("quantity"),
                line.get("line_total"),
            ]
        )
    sheet.append([])
    sheet.append(["Total", quote["pricing"].get("client_price")])
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{quote["quote_id"]}-v{quote["version"]}.xlsx"'
        },
    )


def _pdf(quote: dict):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table

    stream = io.BytesIO()
    document = SimpleDocTemplate(stream, pagesize=A4)
    styles = getSampleStyleSheet()
    rows = [["Part", "Qty", "Total"]] + [
        [
            str(line.get("part_name", "")),
            str(line.get("quantity", "")),
            str(line.get("line_total", "")),
        ]
        for line in quote["pricing"].get("line_items", [])
    ]
    document.build(
        [
            Paragraph(
                f"PartsOps Quote {quote['quote_id']} v{quote['version']}",
                styles["Title"],
            ),
            Paragraph(f"Valid until: {quote['valid_until']}", styles["Normal"]),
            Spacer(1, 12),
            Table(rows),
            Spacer(1, 12),
            Paragraph(
                f"Total: {quote['pricing'].get('client_price', '')}", styles["Heading2"]
            ),
        ]
    )
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{quote["quote_id"]}-v{quote["version"]}.pdf"'
        },
    )
