
from datetime import datetime, timezone
from pathlib import Path
import struct
import zlib

import pytest
from sqlmodel import Session, SQLModel, create_engine, select

from event_store import verify_event_chain
from models import (ClientApproval, ContractExceptionRecord, ContractExport, ContractPosition,
                    ContractWorkflowEvent, ContractWorkflowState, PartRequest, PriceEvidence, RequestState)
from app.automation.context import AutomationContext
from app.automation.runner import run_job
from app.automation.registry import get_job
from services.contract_crawler_adapter import normalize_crawler_rows, normalize_uploaded_crawler_payload
from services.contract_operations import (approve_client_export, approve_contract, archive_contract, authorize_purchase,
                                          collect_evidence, create_contract, evaluate_policy,
                                          export_contract, get_control_plane, advance_contract_workflow,
                                          list_contract_exceptions, register_analog_candidate,
                                          record_purchase, register_oem_candidate, review_position,
                                          update_contract_exception, verify_receipt)


@pytest.fixture
def db():
    engine = create_engine("sqlite://")
    SQLModel.metadata.create_all(engine)
    with Session(engine) as session:
        yield session


def row(part, source, price, screenshot):
    return {"part_number": part, "source": source, "price": price,
            "source_url": f"https://{source}/price/{part}",
            "captured_at": datetime.now(timezone.utc).isoformat(),
            "screenshot_ref": str(screenshot)}


def write_png(path: Path, width: int = 640, height: int = 360):
    def chunk(kind: bytes, payload: bytes) -> bytes:
        return struct.pack(">I", len(payload)) + kind + payload + struct.pack(">I", zlib.crc32(kind + payload) & 0xFFFFFFFF)

    scanline = b"\x00" + (b"\xff\xff\xff" * width)
    raw = scanline * height
    path.write_bytes(
        b"\x89PNG\r\n\x1a\n"
        + chunk(b"IHDR", struct.pack(">IIBBBBB", width, height, 8, 2, 0, 0, 0))
        + chunk(b"IDAT", zlib.compress(raw))
        + chunk(b"IEND", b"")
    )


def seed_verified_candidates(db, request_id, tenant_id="tenant-a", oem_number="OC90"):
    position = db.exec(select(ContractPosition).where(ContractPosition.request_id == request_id)).one()
    oem = register_oem_candidate(db, request_id, tenant_id, position.position_id, {
        "oem_number": oem_number,
        "manufacturer": "BMW",
        "source": "vin_oem_catalog",
        "compatibility_evidence": [
            {"evidence_type": "vin_oem_catalog", "source": "BMW ETK"},
            {"evidence_type": "official_brand_catalog", "source": "BMW catalog"},
            {"evidence_type": "cross_reference", "source": "validated cross"},
        ],
    }, "operator")
    analog = register_analog_candidate(db, request_id, tenant_id, position.position_id, {
        "article": oem_number,
        "brand": "MANN",
        "source": "tecdoc",
        "oem_candidate_id": oem["candidate_id"],
        "independent_confirmations": 2,
        "compatibility_evidence": [
            {"evidence_type": "vin_oem_catalog", "source": "BMW ETK"},
            {"evidence_type": "official_brand_catalog", "source": "MANN catalog"},
            {"evidence_type": "tecdoc", "source": "TecDoc"},
            {"evidence_type": "cross_reference", "source": "validated cross"},
            {"evidence_type": "spec_match", "source": "dimensions"},
        ],
    }, "operator")
    assert oem["verification_status"] == "verified"
    assert analog["manual_review_status"] == "approved"
    return position


def test_contract_evidence_policy_review_and_export(db, tmp_path):
    created = create_contract(db, "tenant-a", [{"part_number": "OC90", "description": "filter", "quantity": 2}], "op")
    request_id = created["request_id"]
    seed_verified_candidates(db, request_id)
    screenshot = tmp_path / "price.png"
    write_png(screenshot)
    collected = collect_evidence(db, request_id, "tenant-a", [row("OC90", "exist.ru", 250, screenshot), row("OC90", "autodoc.ru", 275, screenshot), row("OC90", "rossko.ru", 260, screenshot)], "crawler")
    assert collected["evidence_created"] == 3
    evaluated = evaluate_policy(db, request_id, "tenant-a", "policy")
    assert evaluated["needs_review"] is False
    approve_contract(db, request_id, "tenant-a", "op")
    exported = export_contract(db, request_id, "tenant-a", "op")
    assert exported["lines"][0]["source_url"].startswith("https://")
    assert exported["lines"][0]["screenshot_ref"]
    assert Path(exported["lines"][0]["screenshot_ref"]).is_file()
    assert Path(exported["documents"]["internal_registry"]["path"]).is_file()
    assert Path(exported["documents"]["client_document"]["path"]).is_file()
    assert verify_event_chain(request_id, db, "tenant-a")["valid"]


def test_export_generates_internal_and_client_xlsx_from_same_registry(db, tmp_path):
    from openpyxl import load_workbook

    created = create_contract(db, "tenant-a", [{"part_number": "OC90", "description": "filter", "quantity": 2}], "op")
    request_id = created["request_id"]
    seed_verified_candidates(db, request_id)
    screenshot = tmp_path / "price.png"
    write_png(screenshot)
    collect_evidence(db, request_id, "tenant-a", [
        row("OC90", "exist.ru", 250, screenshot),
        row("OC90", "autodoc.ru", 275, screenshot),
        row("OC90", "rossko.ru", 260, screenshot),
    ], "crawler")
    evaluate_policy(db, request_id, "tenant-a", "policy")
    approve_contract(db, request_id, "tenant-a", "op")
    exported = export_contract(db, request_id, "tenant-a", "op")
    saved = db.exec(select(ContractExport).where(ContractExport.export_id == exported["export_id"])).one()

    assert saved.internal_registry_sha256 == exported["documents"]["internal_registry"]["sha256"]
    assert saved.client_document_sha256 == exported["documents"]["client_document"]["sha256"]
    internal_wb = load_workbook(exported["documents"]["internal_registry"]["path"], read_only=True, data_only=True)
    client_wb = load_workbook(exported["documents"]["client_document"]["path"], read_only=True, data_only=True)
    assert internal_wb["Internal Registry"][1] != client_wb["Client Form"][1]
    assert internal_wb["Export Metadata"]["B3"].value == exported["registry_hash"]
    assert client_wb["Export Metadata"]["B3"].value == exported["registry_hash"]


def test_incomplete_evidence_is_rejected_and_tenant_isolation(db, tmp_path):
    created = create_contract(db, "tenant-a", [{"part_number": "X", "quantity": 1}], "op")
    with pytest.raises(Exception):
        collect_evidence(db, created["request_id"], "tenant-a", [{"part_number": "X", "source": "exist.ru", "price": 1, "source_url": "https://exist.ru/x", "captured_at": datetime.now(timezone.utc).isoformat()}], "crawler")
    with pytest.raises(Exception):
        evaluate_policy(db, created["request_id"], "tenant-b", "policy")


def test_policy_routes_price_spread_to_review(db, tmp_path):
    created = create_contract(db, "tenant-a", [{"part_number": "X", "quantity": 1}], "op")
    seed_verified_candidates(db, created["request_id"], oem_number="X")
    screenshot = tmp_path / "price.png"
    write_png(screenshot)
    collect_evidence(db, created["request_id"], "tenant-a", [
        row("X", "exist.ru", 100, screenshot),
        row("X", "autodoc.ru", 110, screenshot),
        row("X", "rossko.ru", 300, screenshot),
    ], "crawler")
    evaluated = evaluate_policy(db, created["request_id"], "tenant-a", "policy")
    assert evaluated["needs_review"] is True
    position = db.exec(select(ContractPosition).where(ContractPosition.request_id == created["request_id"])).one()
    evidence = db.exec(select(PriceEvidence).where(PriceEvidence.position_id == position.position_id)).first()
    reviewed = review_position(db, created["request_id"], "tenant-a", position.position_id, evidence.evidence_id, "operator", "selected after review")
    assert reviewed["status"] == RequestState.READY_FOR_APPROVAL
    approve_contract(db, created["request_id"], "tenant-a", "operator")
    assert export_contract(db, created["request_id"], "tenant-a", "operator")["lines"]


def test_contract_control_plane_tracks_requirements_gaps_export_and_purchase_lock(db, tmp_path):
    created = create_contract(db, "tenant-a", [{"part_number": "OC90", "description": "filter", "quantity": 2}], "op")
    request_id = created["request_id"]
    initial = get_control_plane(db, request_id, "tenant-a")
    assert len(initial["requirements"]) >= 10
    assert any(req["clause"] == "AUD-REQ-001" for req in initial["requirements"])
    assert any(gap["status"] == "open" for gap in initial["gaps"])

    screenshot = tmp_path / "price.png"
    write_png(screenshot)
    seed_verified_candidates(db, request_id)
    collect_evidence(db, request_id, "tenant-a", [
        row("OC90", "exist.ru", 250, screenshot),
        row("OC90", "autodoc.ru", 275, screenshot),
        row("OC90", "rossko.ru", 260, screenshot),
    ], "crawler")
    evaluate_policy(db, request_id, "tenant-a", "policy")
    position = db.exec(select(ContractPosition).where(ContractPosition.request_id == request_id)).one()
    assert position.calculation_json
    assert '"coefficient": 0.99' in position.calculation_json

    approve_contract(db, request_id, "tenant-a", "operator")
    exported = export_contract(db, request_id, "tenant-a", "operator")
    assert exported["registry_hash"]
    assert exported["lines"][0]["unit_price"] == 247.5
    assert exported["lines"][0]["total"] == 495.0

    with pytest.raises(Exception):
        authorize_purchase(db, request_id, "tenant-a", "missing-approval", "operator")
    assert db.exec(select(ContractExceptionRecord).where(
        ContractExceptionRecord.request_id == request_id,
        ContractExceptionRecord.code == "PUR-001",
    )).first()

    approved = approve_client_export(db, request_id, "tenant-a", exported["export_id"], "customer", "email://approval")
    authorized = authorize_purchase(db, request_id, "tenant-a", approved["approval_id"], "operator")
    assert authorized["status"] == "authorized"
    assert db.exec(select(ClientApproval).where(ClientApproval.approval_id == approved["approval_id"])).one()
    final_control = get_control_plane(db, request_id, "tenant-a")
    purchase_req = next(req for req in final_control["requirements"] if req["clause"] == "PUR-REQ-001")
    assert purchase_req["coverage_status"] == "Covered"
    assert final_control["metrics"]["quality"]["requirements_covered"] >= 10
    assert final_control["metrics"]["evidence"]["required_source_coverage_percent"] == 100.0
    assert final_control["metrics"]["evidence"]["screenshot_coverage_percent"] == 100.0
    assert final_control["metrics"]["cost"]["contract_total"] == 495.0
    assert final_control["metrics"]["process"]["purchase_locked"] is False


def test_purchase_receipt_archive_workflow_tail_is_enforced_and_audited(db, tmp_path):
    created = create_contract(db, "tenant-a", [{"part_number": "OC90", "description": "filter", "quantity": 2}], "op")
    request_id = created["request_id"]
    seed_verified_candidates(db, request_id)
    screenshot = tmp_path / "price.png"
    write_png(screenshot)
    collect_evidence(db, request_id, "tenant-a", [
        row("OC90", "exist.ru", 250, screenshot),
        row("OC90", "autodoc.ru", 275, screenshot),
        row("OC90", "rossko.ru", 260, screenshot),
    ], "crawler")
    evaluate_policy(db, request_id, "tenant-a", "policy")
    approve_contract(db, request_id, "tenant-a", "operator")
    exported = export_contract(db, request_id, "tenant-a", "operator")
    approved = approve_client_export(db, request_id, "tenant-a", exported["export_id"], "customer", "email://approval")

    with pytest.raises(Exception):
        record_purchase(db, request_id, "tenant-a", "missing-auth", "SUP-1", "operator")

    authorized = authorize_purchase(db, request_id, "tenant-a", approved["approval_id"], "operator")
    purchase = record_purchase(db, request_id, "tenant-a", authorized["authorization_id"],
                               "SUP-EXIST", "buyer", "po://1001")
    control = get_control_plane(db, request_id, "tenant-a")
    assert control["workflow"]["current_stage"] == "24_PURCHASED"
    assert control["purchases"][0]["amount_total"] == 495.0

    receipt = verify_receipt(db, request_id, "tenant-a", purchase["purchase_id"],
                             "warehouse", "receipt://1001", 2)
    archive = archive_contract(db, request_id, "tenant-a", receipt["receipt_id"],
                               "operator", "archive://contract-package")
    final_control = get_control_plane(db, request_id, "tenant-a")
    assert archive["status"] == "archived"
    assert final_control["workflow"]["current_stage"] == "26_ARCHIVED"
    assert final_control["receipt_verifications"][0]["received_quantity"] == 2
    assert final_control["archives"][0]["registry_hash"] == exported["registry_hash"]
    assert final_control["metrics"]["process"]["purchases"] == 1
    assert final_control["metrics"]["process"]["receipt_verifications"] == 1
    assert final_control["metrics"]["process"]["archives"] == 1
    assert verify_event_chain(request_id, db, "tenant-a")["valid"]


def test_contract_exception_lifecycle_is_audited(db):
    created = create_contract(db, "tenant-a", [{"part_number": "X", "quantity": 1}], "op")
    request_id = created["request_id"]
    with pytest.raises(Exception):
        authorize_purchase(db, request_id, "tenant-a", "missing-approval", "operator")

    listed = list_contract_exceptions(db, request_id, "tenant-a")
    exception_id = listed["exceptions"][0]["exception_id"]
    assigned = update_contract_exception(db, request_id, "tenant-a", exception_id, "assign",
                                         "lead", owner="commercial-reviewer")
    assert assigned["owner"] == "commercial-reviewer"
    retried = update_contract_exception(db, request_id, "tenant-a", exception_id, "retry", "lead")
    assert retried["retry_count"] == 1
    escalated = update_contract_exception(db, request_id, "tenant-a", exception_id, "escalate",
                                          "lead", owner="head-of-purchase")
    assert escalated["severity"] == "CRITICAL"
    assert escalated["escalation_due_at"]
    resolved = update_contract_exception(db, request_id, "tenant-a", exception_id, "resolve",
                                         "head-of-purchase", resolution="Client approval attached later",
                                         evidence_ref="email://approval")
    assert resolved["status"] == "resolved"
    assert get_control_plane(db, request_id, "tenant-a")["exceptions"][0]["status"] == "resolved"
    assert verify_event_chain(request_id, db, "tenant-a")["valid"]


def test_stale_price_evidence_blocks_policy_and_export(db, tmp_path):
    created = create_contract(db, "tenant-a", [{"part_number": "X", "quantity": 1}], "op")
    seed_verified_candidates(db, created["request_id"], oem_number="X")
    screenshot = tmp_path / "price.png"
    write_png(screenshot)
    collect_evidence(db, created["request_id"], "tenant-a", [
        {**row("X", "exist.ru", 100, screenshot), "freshness_ttl_hours": 1},
        {**row("X", "autodoc.ru", 101, screenshot), "freshness_ttl_hours": 1},
        {**row("X", "rossko.ru", 102, screenshot), "freshness_ttl_hours": 1},
    ], "crawler")
    for evidence in db.exec(select(PriceEvidence).where(PriceEvidence.request_id == created["request_id"])).all():
        evidence.expires_at = datetime(2000, 1, 1)
        db.add(evidence)
    db.commit()
    evaluated = evaluate_policy(db, created["request_id"], "tenant-a", "policy")
    assert evaluated["needs_review"] is True
    control = get_control_plane(db, created["request_id"], "tenant-a")
    evidence_req = next(req for req in control["requirements"] if req["clause"] == "EVD-REQ-001")
    assert evidence_req["coverage_status"] == "Partial"


def test_price_collection_requires_verified_oem_and_approved_analog(db, tmp_path):
    created = create_contract(db, "tenant-a", [{"part_number": "X", "quantity": 1}], "op")
    screenshot = tmp_path / "price.png"
    write_png(screenshot)
    with pytest.raises(Exception):
        collect_evidence(db, created["request_id"], "tenant-a", [row("X", "exist.ru", 100, screenshot)], "crawler")

    position = db.exec(select(ContractPosition).where(ContractPosition.request_id == created["request_id"])).one()
    low_oem = register_oem_candidate(db, created["request_id"], "tenant-a", position.position_id, {
        "oem_number": "X",
        "source": "weak_catalog",
        "compatibility_evidence": [{"evidence_type": "cross_reference", "source": "weak"}],
    }, "operator")
    assert low_oem["verification_status"] == "needs_review"

    seed_verified_candidates(db, created["request_id"], oem_number="X")
    collected = collect_evidence(db, created["request_id"], "tenant-a", [row("X", "exist.ru", 100, screenshot)], "crawler")
    assert collected["evidence_created"] == 1


def test_unreadable_screenshot_is_rejected_before_price_evidence(db, tmp_path):
    created = create_contract(db, "tenant-a", [{"part_number": "X", "quantity": 1}], "op")
    seed_verified_candidates(db, created["request_id"], oem_number="X")
    screenshot = tmp_path / "broken.png"
    screenshot.write_bytes(b"\x89PNG\r\n\x1a\ncontract-evidence")
    with pytest.raises(Exception):
        collect_evidence(db, created["request_id"], "tenant-a", [row("X", "exist.ru", 100, screenshot)], "crawler")
    assert not db.exec(select(PriceEvidence).where(PriceEvidence.request_id == created["request_id"])).all()


def test_crawler_adapter_normalizes_deduplicates_and_skips_invalid_rows(tmp_path):
    screenshot = tmp_path / "shot.png"
    write_png(screenshot)
    rows, stats = normalize_crawler_rows([
        {"site": "exist.ru", "article": "X", "price": "1 250,50 ₽", "url": "https://exist.ru/price/X",
         "capturedAt": "2026-07-24T10:00:00+00:00", "screenshot": screenshot.name, "stock_qty": 4},
        {"site": "exist.ru", "article": "X", "price": "1 250,50 ₽", "url": "https://exist.ru/price/X",
         "capturedAt": "2026-07-24T10:00:00+00:00", "screenshot": screenshot.name},
        {"site": "unknown", "article": "X", "price": "10", "url": "https://example.test/x",
         "capturedAt": "2026-07-24T10:00:00+00:00", "screenshot": screenshot.name},
        {"site": "rossko.ru", "article": "X", "price": "——", "url": "https://rossko.ru/x",
         "capturedAt": "2026-07-24T10:00:00+00:00", "screenshot": screenshot.name},
    ], tmp_path)

    assert stats == {"input": 4, "normalized": 1, "skipped": 2, "duplicates": 1}
    assert rows[0]["source"] == "exist.ru"
    assert rows[0]["price"] == 1250.50
    assert rows[0]["screenshot_ref"] == str(screenshot)
    assert rows[0]["available_quantity"] == 4


def test_uploaded_crawler_payload_requires_absolute_screenshot_paths():
    rows, stats = normalize_uploaded_crawler_payload(b"""{"items":[{
        "site":"exist.ru",
        "article":"X",
        "price":"10",
        "url":"https://exist.ru/x",
        "capturedAt":"2026-07-24T10:00:00+00:00",
        "screenshot":"relative.png"
    }]}""")
    assert rows == []
    assert stats["skipped"] == 1


def test_workflow_v2_blocks_forbidden_transitions_and_tracks_history(db, tmp_path):
    created = create_contract(db, "tenant-a", [{"part_number": "X", "quantity": 1}], "op")
    request_id = created["request_id"]
    workflow = db.exec(select(ContractWorkflowState).where(ContractWorkflowState.request_id == request_id)).one()
    assert workflow.current_stage == "04_NEW_POSITION"

    with pytest.raises(Exception):
        advance_contract_workflow(db, request_id, "tenant-a", "11_PRICES_COLLECTED", "operator", "too early")
    rejected = db.exec(select(ContractWorkflowEvent).where(
        ContractWorkflowEvent.request_id == request_id,
        ContractWorkflowEvent.allowed == False,  # noqa: E712
    )).first()
    assert rejected
    assert "MKT-001" in rejected.violations_json

    seed_verified_candidates(db, request_id, oem_number="X")
    workflow = db.exec(select(ContractWorkflowState).where(ContractWorkflowState.request_id == request_id)).one()
    assert workflow.current_stage == "10_ANALOGS_VERIFIED"

    screenshot = tmp_path / "price.png"
    write_png(screenshot)
    collect_evidence(db, request_id, "tenant-a", [
        row("X", "exist.ru", 100, screenshot),
        row("X", "autodoc.ru", 101, screenshot),
        row("X", "rossko.ru", 102, screenshot),
    ], "crawler")
    evaluate_policy(db, request_id, "tenant-a", "policy")
    workflow = db.exec(select(ContractWorkflowState).where(ContractWorkflowState.request_id == request_id)).one()
    assert workflow.current_stage == "15_AGENT_QA_PASSED"

    control = get_control_plane(db, request_id, "tenant-a")
    assert control["workflow"]["current_stage"] == "15_AGENT_QA_PASSED"
    assert any(event["allowed"] is False for event in control["workflow"]["events"])


def test_contract_orchestrate_job_executes_agent_roles_without_bypassing_human_gate(db, tmp_path):
    assert get_job("contract_orchestrate")
    created = create_contract(db, "tenant-a", [{"part_number": "X", "quantity": 1}], "op")
    request_id = created["request_id"]
    seed_verified_candidates(db, request_id, oem_number="X")
    screenshot = tmp_path / "price.png"
    write_png(screenshot)
    collect_evidence(db, request_id, "tenant-a", [
        row("X", "exist.ru", 100, screenshot),
        row("X", "autodoc.ru", 101, screenshot),
        row("X", "rossko.ru", 102, screenshot),
    ], "crawler")

    result = run_job(db, "contract_orchestrate", AutomationContext(
        tenant_id="tenant-a",
        request_id=request_id,
        actor_id="contract-agent",
        correlation_id="CORR-CONTRACT-ORCH",
    ))

    assert result["ok"] is True
    assert result["result"]["processed"] == 1
    assert result["result"]["blocked_roles"] == 0
    assert any(role["role"] == "pricing_policy_agent" and role["status"] == "completed"
               for role in result["result"]["requests"][0]["roles"])
    req = db.exec(select(PartRequest).where(PartRequest.request_id == request_id)).one()
    assert req.status == RequestState.READY_FOR_APPROVAL
    assert get_control_plane(db, request_id, "tenant-a")["workflow"]["current_stage"] == "15_AGENT_QA_PASSED"
    assert verify_event_chain(request_id, db, "tenant-a")["valid"]
